import pandas as pd
from openpyxl import load_workbook

# to generate random password for students
import string    
import random
S = 10  

specialChars = ['!','@','#','$','%','&','*',]

def givePassword():
    words = ''.join(random.choices(string.ascii_uppercase, k = 3)) + ''.join(random.choices(string.ascii_lowercase, k = 3)) + ''.join(random.choices(specialChars, k = 2)) + ''.join(random.choices(string.digits, k = 2))
    return words

def setpasswords(filepath):
    workbook = load_workbook(filename=filepath)
    wb = workbook.active
    count = wb.max_row
    wb['F1'] = 'password'
    for i in range(2,count+1):
        wb['F'+str(i)] = givePassword()

    workbook.save(filepath)


def extractData(file):
    setpasswords(file)
    data = pd.read_excel(file)
    print("length is: ",len(data['rollNo']))
    result = {}
    for r,f,l,fn,m,p in zip(data['rollNo'],data['first name'],data['last name'],data['father name'], data['mother name'], data['password']):
        temp = {}
        temp['First Name'],temp['Last name'],temp['Father Name'],temp['Mother Name'],temp['password'] = f,l,fn,m,p
        result[r] = temp

    return result


if '__main__' == __name__:
    print(givePassword())
    # print(extractData('rnos.xlsx'))